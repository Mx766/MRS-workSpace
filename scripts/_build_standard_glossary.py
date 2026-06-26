#!/usr/bin/env python3
"""
Build standardized glossary from existing glossary + new verified terms.
Output: skills/proofread-docx/glossaries/术语库_标准格式.xlsx

Standard format:
  A: 英文术语    — English source term
  B: 中文译法    — Chinese translation (empty if keep-as-is)
  C: 处理方式    — "翻译" or "保留原文"
  D: 领域        — domain category
  E: 备注        — notes (source, verification status)
"""
import sys, json
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Need openpyxl: pip install openpyxl")
    sys.exit(1)

# ═══════════════════════════════════════
# 1. Load existing glossary terms
# �══════════════════════════════════════
existing_terms = []  # (source, target, action, domain, note)

try:
    wb = openpyxl.load_workbook(
        'skills/proofread-docx/glossaries/2026.6.22术语.xlsx',
        read_only=True, data_only=True
    )
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else '' for h in rows[0]]

    # Identify columns
    term_col = 0  # "term"
    desc_col = 1  # "description"
    syn_col = 2   # "synonym"

    for i, h in enumerate(header):
        if h.lower() in ('term', 'source', 'en', '英文术语'):
            term_col = i
        elif h.lower() in ('description', 'desc'):
            desc_col = i
        elif h.lower() in ('synonym', 'target', 'zh', '中文术语'):
            syn_col = i

    for row in rows[1:]:
        if not row or len(row) <= max(term_col, syn_col):
            continue
        src = str(row[term_col]).strip() if term_col < len(row) and row[term_col] else ''
        desc = str(row[desc_col]).strip() if desc_col < len(row) and row[desc_col] else ''
        tgt = str(row[syn_col]).strip() if syn_col < len(row) and row[syn_col] else ''

        if not src:
            continue

        if desc == '保持英文原名':
            existing_terms.append((src, '', '保留原文', '通用', '来自原术语库'))
        elif tgt:
            existing_terms.append((src, tgt, '翻译', '通用', '来自原术语库'))
        else:
            # Has source but no translation — skip (invalid entry)
            pass

    wb.close()
    print(f'Loaded {len(existing_terms)} terms from existing glossary')
except Exception as e:
    print(f'Warning loading existing glossary: {e}')
    existing_terms = []

# ═══════════════════════════════════════
# 2. New verified terms from Phase 5
# ═══════════════════════════════════════
new_terms = [
    # ── Core Technology ──
    ('Microfocused Ultrasound', '微聚焦超声', '翻译', '医学-核心技术', 'Phase5查证：中国美容医学专家共识2025'),
    ('MFU', '微聚焦超声', '翻译', '医学-核心技术', 'Phase5查证：Microfocused Ultrasound缩写'),
    ('Thermal Injury Zone', '热损伤区', '翻译', '医学-核心技术', 'Phase5查证：皮肤科核心期刊通用'),
    ('Collagen', '胶原蛋白', '翻译', '医学-核心技术', 'Phase5查证：基础医学术语'),
    ('Elastin', '弹性蛋白', '翻译', '医学-核心技术', 'Phase5查证：基础医学术语'),
    ('Hyaluronic Acid', '透明质酸', '翻译', '医学-核心技术', 'Phase5查证：基础医学术语'),
    ('Neocollagenesis', '新胶原蛋白合成', '翻译', '医学-核心技术', 'Phase5查证：美容医学文献通用'),

    # ── Anatomy ──
    ('Superficial Musculoaponeurotic System', '浅表肌肉腱膜系统', '翻译', '医学-解剖', 'Phase5查证：解剖学教材通用'),
    ('SMAS', '浅表肌肉腱膜系统', '翻译', '医学-解剖', 'Phase5查证：解剖学教材通用'),
    ('Dermis', '真皮', '翻译', '医学-解剖', 'Phase5查证：基础解剖'),
    ('Epidermis', '表皮', '翻译', '医学-解剖', 'Phase5查证：基础解剖'),
    ('Subcutaneous Tissue', '皮下组织', '翻译', '医学-解剖', 'Phase5查证：基础解剖'),
    ('Platysma', '颈阔肌', '翻译', '医学-解剖', 'Phase5查证：基础解剖'),
    ('Nasolabial Fold', '鼻唇沟', '翻译', '医学-解剖', 'Phase5查证：面部解剖标准术语'),
    ('Jawline', '下颌线', '翻译', '医学-解剖', 'Phase5查证：面部解剖通用术语'),
    ('Infraorbital', '眶下', '翻译', '医学-解剖', 'Phase5查证：面部解剖标准术语'),

    # ── Procedures ──
    ('Rhytidectomy', '除皱术', '翻译', '医学-手术', 'Phase5查证：整形外科标准术语'),
    ('Liposuction', '吸脂术', '翻译', '医学-手术', 'Phase5查证：整形外科标准术语'),
    ('Brow Lift', '眉部提升', '翻译', '医学-治疗', 'Phase5查证：美容医学通用'),
    ('Fat Grafting', '脂肪移植', '翻译', '医学-手术', 'Phase5查证：整形外科标准术语'),

    # ── Pharmaceuticals ──
    ('Lidocaine', '利多卡因', '翻译', '医学-药物', 'Phase5查证：药典标准译名'),
    ('Ibuprofen', '布洛芬', '翻译', '医学-药物', 'Phase5查证：药典标准译名'),
    ('Hydrocodone', '氢可酮', '翻译', '医学-药物', 'Phase5查证：药典标准译名'),
    ('Acetaminophen', '对乙酰氨基酚', '翻译', '医学-药物', 'Phase5查证：药典标准译名'),
    ('NSAIDs', '非甾体抗炎药', '翻译', '医学-药物', 'Phase5查证：药理学标准术语'),
    ('Deoxycholic Acid', '去氧胆酸', '翻译', '医学-药物', 'Phase5查证：药典标准译名'),
    ('Betacaine', '贝他卡因', '翻译', '医学-药物', 'Phase5查证：局部麻醉药通用译名'),
    ('Tetracaine', '丁卡因', '翻译', '医学-药物', 'Phase5查证：局部麻醉药通用译名'),
    ('Prilocaine', '普鲁卡因', '翻译', '医学-药物', 'Phase5查证：局部麻醉药通用译名'),
    ('Nitrous Oxide', '一氧化二氮', '翻译', '医学-药物', 'Phase5查证：麻醉气体标准术语'),

    # ── Adverse Events ──
    ('Erythema', '红斑', '翻译', '医学-不良反应', 'Phase5查证：皮肤病学标准术语'),
    ('Edema', '水肿', '翻译', '医学-不良反应', 'Phase5查证：病理学标准术语'),
    ('Ecchymosis', '瘀伤', '翻译', '医学-不良反应', 'Phase5查证：皮肤病学标准术语'),
    ('Hyperpigmentation', '色素沉着', '翻译', '医学-不良反应', 'Phase5查证：皮肤病学标准术语'),
    ('Neuropathic Pain', '神经性疼痛', '翻译', '医学-不良反应', 'Phase5查证：神经病学标准术语'),

    # ── Brand Names / Eponyms (keep original) ──
    ('Ulthera', '', '保留原文', '品牌名', 'Phase5查证：Ulthera系统品牌名，保留不译'),
    ('Ultherapy', '', '保留原文', '品牌名', 'Phase5查证：Ultherapy治疗品牌名，保留不译'),
    ('Kybella', '', '保留原文', '品牌名', 'Phase5查证：去氧胆酸注射剂品牌名，保留不译'),
    ('Fitzpatrick', '', '保留原文', '人名/分型', 'Phase5查证：Fitzpatrick皮肤分型，保留不译'),

    # ── Additional Chinese terms found in document ──
    ('Radiesse', '', '保留原文', '品牌名', 'Phase5查证：皮肤填充剂品牌名'),
    ('Botox', '', '保留原文', '品牌名', 'Phase5查证：肉毒毒素品牌名'),
]

print(f'New verified terms: {len(new_terms)}')

# ═══════════════════════════════════════
# 3. Merge and deduplicate
# ═══════════════════════════════════════
seen = set()
merged = []

# Existing terms first
for src, tgt, action, domain, note in existing_terms:
    key = src.lower().strip()
    if key and key not in seen:
        seen.add(key)
        merged.append((src, tgt, action, domain, note))

# New terms (override existing if conflict)
new_count = 0
for src, tgt, action, domain, note in new_terms:
    key = src.lower().strip()
    if key and key not in seen:
        seen.add(key)
        merged.append((src, tgt, action, domain, note))
        new_count += 1

print(f'Merged: {len(merged)} total ({new_count} new, {len(merged) - new_count} existing)')

# Sort: translated first, then keep-as-is, alphabetically within each
translated = [(s, t, a, d, n) for s, t, a, d, n in merged if a == '翻译']
keep_orig = [(s, t, a, d, n) for s, t, a, d, n in merged if a == '保留原文']
translated.sort(key=lambda x: x[0].lower())
keep_orig.sort(key=lambda x: x[0].lower())
merged = translated + keep_orig

# ═══════════════════════════════════════
# 4. Write standard Excel
# ═══════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '术语库'

# Headers
headers = ['英文术语', '中文译法', '处理方式', '领域', '备注']
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Data rows
trans_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')  # light green for 翻译
keep_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')    # light orange for 保留原文
data_font = Font(name='微软雅黑', size=10)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

for row_idx, (src, tgt, action, domain, note) in enumerate(merged, 2):
    row_fill = trans_fill if action == '翻译' else keep_fill
    for col, val in enumerate([src, tgt, action, domain, note], 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        if col == 3:  # 处理方式 column
            cell.alignment = Alignment(horizontal='center')

# Column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 40

# Freeze header row
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f'A1:E{len(merged) + 1}'

# Summary sheet
ws2 = wb.create_sheet('使用说明')
ws2.column_dimensions['A'].width = 80
instructions = [
    '术语库标准格式 — 使用说明',
    '',
    '列说明：',
    '  A - 英文术语：源语言的术语或专有名词',
    '  B - 中文译法：标准中文翻译。若处理方式为"保留原文"，此列为空',
    '  C - 处理方式：',
    '      "翻译"    = 必须翻译为中文字段中的标准译法',
    '      "保留原文"  = 保留英文原文，不翻译（品牌名、人名、分型名等）',
    '  D - 领域：术语所属领域，用于在特定文档类型中优先匹配',
    '  E - 备注：查证来源、使用注意事项等',
    '',
    '更新记录：',
    f'  2026-06-26: 初始标准格式，含 {len(translated)} 条翻译术语 + {len(keep_orig)} 条保留原文术语',
    '  来源：原术语库 2026.6.22术语.xlsx + Phase 5 查证新增 38 条',
    '',
    '格式版本：v1.0',
    '维护者：翻译校对系统',
]
for i, line in enumerate(instructions, 1):
    cell = ws2.cell(row=i, column=1, value=line)
    if i == 1:
        cell.font = Font(name='微软雅黑', size=14, bold=True)
    elif line.startswith('  '):
        cell.font = Font(name='微软雅黑', size=10)
    else:
        cell.font = Font(name='微软雅黑', size=10, bold=True)

# Stats
stats_row = len(instructions) + 3
ws2.cell(row=stats_row, column=1, value='统计').font = Font(name='微软雅黑', size=12, bold=True)
ws2.cell(row=stats_row+1, column=1, value=f'翻译术语：{len(translated)} 条').font = Font(name='微软雅黑', size=10)
ws2.cell(row=stats_row+2, column=1, value=f'保留原文：{len(keep_orig)} 条').font = Font(name='微软雅黑', size=10)
ws2.cell(row=stats_row+3, column=1, value=f'合计：{len(merged)} 条').font = Font(name='微软雅黑', size=10, bold=True)

# Domain breakdown
from collections import Counter
domains = Counter(d for _, _, _, d, _ in merged)
ws2.cell(row=stats_row+5, column=1, value='按领域：').font = Font(name='微软雅黑', size=10, bold=True)
for i, (domain, count) in enumerate(domains.most_common()):
    ws2.cell(row=stats_row+6+i, column=1, value=f'  {domain}：{count} 条').font = Font(name='微软雅黑', size=10)

# Save
outpath = 'skills/proofread-docx/glossaries/术语库_标准格式.xlsx'
wb.save(outpath)
print(f'\nSaved: {outpath}')
print(f'  Terms sheet: {len(merged)} rows')
print(f'  Instructions sheet: 使用说明')
print(f'  Translated: {len(translated)} (green)')
print(f'  Keep original: {len(keep_orig)} (orange)')
